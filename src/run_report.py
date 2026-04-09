
# src/run_report.py
# 半量化交易系統整合版
# ─────────────────────────────────────────────────────────────────────────────
# 交易訊號邏輯（依 Word 文件「半量化交易系統」）：
#   進場價  = MAX(今日收盤價, 20日高點) × 1.01
#   停損價  = 進場價 × 0.92
#   停利1   = 進場價 × 1.20
#   停利2   = 進場價 × 1.35
#   移動停損 = 10日均線
#   是否突破20日高 = 今日最高價 > 20日高點
#   是否放量1.5倍  = 今日成交量 > 5日均量 × 1.5
#   大盤站上月線   = 大盤指數 > 大盤20日均線
#   綜合判斷       = 突破 + 放量 + 大盤 + 外資淨買超 → 🔥可買 / 觀察
# ─────────────────────────────────────────────────────────────────────────────

import os
import re
import json
import time
import random
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from zoneinfo import ZoneInfo
import contextlib
import io

import pandas as pd
import requests
from bs4 import BeautifulSoup
import yfinance as yf

from config import LEGEND_BROKERS

try:
    from market_data import fetch_all_market_data
    HAS_MARKET_DATA = True
except ImportError:
    HAS_MARKET_DATA = False
    print("[WARN] market_data 模組未找到，跳過大盤籌碼抓取")

TZ = ZoneInfo("Asia/Taipei")
DEFAULT_DAYS = 5
BASE_URL = "https://fubon-ebrokerdj.fbs.com.tw/z/zg/zgb/zgb0.djhtm"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

NO_DATA_TEXT = "無此券商分點交易資料"

# ── yfinance 共用 Session（curl_cffi 連線池）─────────────────────────────────
# yfinance 1.2.x 起改用 curl_cffi 做 TLS 指紋模擬繞過 Cloudflare 防護。
# 傳入標準 requests.Session 會直接拋出：
#   "Yahoo API requires curl_cffi session not requests.sessions.Session"
# 必須用 curl_cffi.requests.Session 才能同時獲得：
#   ① 正確的 TLS 指紋（通過 Yahoo 驗證）
#   ② 連線池重用（降低 429 風險、提升速度）
try:
    from curl_cffi.requests import Session as CurlSession
    _YF_SESSION = CurlSession(impersonate="chrome")
    _YF_SESSION_OK = True
except ImportError:
    # curl_cffi 未安裝時退回 None，yfinance 自行管理 session
    _YF_SESSION = None
    _YF_SESSION_OK = False
    print("[WARN] curl_cffi 未安裝，yfinance 將使用預設 session（pip install curl_cffi 可提速）")

# ── 快取 ──────────────────────────────────────────────────────────────────────
TECH_DATA_CACHE: dict = {}   # 個股技術面快取
TAIEX_MA20_CACHE: dict = {}  # 大盤月線快取（只需一份）
_CACHE_LOCK = threading.Lock()  # 多執行緒寫入快取時的互斥鎖

# 並行抓取的 thread 數量。yfinance 對同一 IP 的並發連線有隱性限制，
# 實測 8~12 最穩定；設太高反而觸發 429/rate-limit 導致全部失敗。
TECH_WORKERS = int(os.getenv("TECH_WORKERS", "10"))


# ─────────────────────────────────────────────────────────────────────────────
#  工具函式
# ─────────────────────────────────────────────────────────────────────────────
def ensure_output_dir():
    os.makedirs("output", exist_ok=True)


def now_taipei():
    return datetime.now(TZ)


def safe_int(s: str) -> int:
    if s is None:
        return 0
    s = s.strip().replace(",", "")
    if s in ("", "--"):
        return 0
    try:
        return int(s)
    except ValueError:
        m = re.search(r"-?\d+", s)
        return int(m.group()) if m else 0


def dump_debug_html(tag: str, html: str):
    if os.getenv("DEBUG_HTML", "0") != "1":
        return
    os.makedirs("output/debug", exist_ok=True)
    path = os.path.join("output", "debug", f"{tag}.html")
    with open(path, "w", encoding="utf-8", errors="ignore") as f:
        f.write(html)


def fetch_html(url: str, timeout: int = 20, retries: int = 5, base_sleep: float = 1.0) -> str:
    headers = {
        "User-Agent": UA,
        "Referer": "https://fubon-ebrokerdj.fbs.com.tw/",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
        "Connection": "keep-alive",
    }
    last_err = None
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, timeout=timeout)
            r.encoding = "big5"
            if r.status_code == 200 and r.text:
                return r.text
            last_err = RuntimeError(f"HTTP {r.status_code}")
        except Exception as e:
            last_err = e
        time.sleep(base_sleep * (2 ** attempt) + random.uniform(0, 0.5))
    raise last_err


def parse_table(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    res = {}
    pat = re.compile(
        r"GenLink2stk\(\s*'AS(\w+)'\s*,\s*'(.+?)'\s*\)"
        r"|GenLink2stk\(\s*\"AS(\w+)\"\s*,\s*\"(.+?)\"\s*\)"
    )
    for tr in soup.find_all("tr"):
        found = None
        for sc in tr.find_all("script"):
            txt = sc.get_text() or ""
            m = pat.search(txt)
            if m:
                found = (m.group(1) or m.group(3), m.group(2) or m.group(4))
                break
        if not found:
            for a in tr.find_all("a"):
                blob = (a.get("href", "") or "") + " " + (a.get("onclick", "") or "")
                m = pat.search(blob)
                if m:
                    found = (m.group(1) or m.group(3), m.group(2) or m.group(4))
                    break
        if not found:
            m = pat.search(str(tr))
            if m:
                found = (m.group(1) or m.group(3), m.group(2) or m.group(4))
        if not found:
            continue
        sid, name = found
        tds = tr.find_all("td", class_=["t3n1", "t3n0"])
        if len(tds) < 3:
            tds = tr.find_all("td")
        if len(tds) < 3:
            continue
        res[sid] = {
            "name": name,
            "buy": tds[0].get_text(strip=True),
            "sell": tds[1].get_text(strip=True),
            "net": safe_int(tds[2].get_text(strip=True)),
        }
    return res


# ─────────────────────────────────────────────────────────────────────────────
#  技術面數據（1個月歷史）
# ─────────────────────────────────────────────────────────────────────────────
def get_stock_tech_data(sid: str) -> dict:
    """
    靜音 yfinance 噪音 + 快取。
    抓取 1 個月歷史數據，計算：
        price    現價（今日收盤）
        high     今日最高價
        low      今日最低價
        vol      今日成交量
        high_20  20 日最高價（突破基準）
        ma_10    10 日均線（移動停損參考）
        vol_5    5 日均量（放量判斷）
    """
    if sid in TECH_DATA_CACHE:
        return TECH_DATA_CACHE[sid]

    buf = io.StringIO()
    data = {
        "price": 0.0, "high": 0.0, "low": 0.0, "vol": 0,
        "high_20": 0.0, "ma_10": 0.0, "vol_5": 0,
    }

    # yfinance 重試常數（可透過環境變數覆寫）
    YF_RETRIES = int(os.getenv("YF_RETRIES",  "3"))    # 最多重試次數
    YF_BACKOFF = float(os.getenv("YF_BACKOFF", "1.5"))  # 指數退避底數（秒）

    # ── 迴圈結構說明 ─────────────────────────────────────────────────────────
    # 原本設計：外層 suffix → 內層 attempt
    # 問題：上櫃股（.TWO）一定在 .TW 拿到空值，舊版每次都 sleep 重試 3 次
    #       才換 suffix，浪費 ≈ 1.5+2.25 = 3.75 秒/每檔上櫃股。
    #
    # 修正後：外層 attempt → 內層 suffix
    # 同一次嘗試內，.TW 和 .TWO 連續各試一次，任一成功立即 return。
    # 兩個 suffix 都失敗（真正的網路問題），才 sleep 進行下一次 attempt。
    # 上櫃股不再有無謂的 sleep，效能大幅提升。
    for attempt in range(1, YF_RETRIES + 1):
        for suffix in [".TW", ".TWO"]:
            try:
                with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
                    # ⚠️ 用 3mo 而非 1mo：
                    #   period="1mo" 是「日曆月」，遇春節/清明連假實際交易日
                    #   可能只有 13~15 天，導致 tail(20) 實際上只有 15 日高點。
                    #   用 3mo（≈ 60 交易日）可確保任何假期情境下
                    #   tail(20)/tail(10)/tail(5) 都是貨真價實的交易日數量。
                    # curl_cffi Session 可用時傳入（連線池 + 正確 TLS 指紋）
                    # 不可用時讓 yfinance 自行建立 session（功能正常，略慢）
                    ticker_kwargs = {"session": _YF_SESSION} if _YF_SESSION_OK else {}
                    hist = yf.Ticker(f"{sid}{suffix}", **ticker_kwargs).history(period="3mo")

                if hist is None or hist.empty:
                    continue

                # ── NaN 防禦 ─────────────────────────────────────────────────
                # 剛上市股、暫停交易股或伺服器資料異常時，欄位可能含有 NaN。
                # NaN 會污染 .max()/.mean() → 使結果也變 NaN → 存入 JSON 後
                # 標準解析器不認識 NaN 直接崩潰。用 .dropna() 先行清除。
                close  = hist["Close"].dropna()
                high   = hist["High"].dropna()
                low    = hist["Low"].dropna()
                volume = hist["Volume"].dropna()

                # 所有關鍵欄位都有至少一個有效值才繼續
                if close.empty or high.empty or volume.empty:
                    continue

                def _safe_float(series, default=0.0) -> float:
                    val = series.iloc[-1] if not series.empty else default
                    import math
                    return default if (val is None or (isinstance(val, float) and math.isnan(val))) else float(val)

                def _safe_int(series, default=0) -> int:
                    val = series.iloc[-1] if not series.empty else default
                    import math
                    return default if (val is None or (isinstance(val, float) and math.isnan(val))) else int(val)

                data["price"]   = _safe_float(close)
                data["high"]    = _safe_float(high)
                data["low"]     = _safe_float(low) if not low.empty else 0.0
                data["vol"]     = _safe_int(volume)
                # tail(N) 對齊實際交易日，dropna 確保計算不被 NaN 污染
                data["high_20"] = float(high.tail(20).max())   if len(high)   >= 1 else 0.0
                data["ma_10"]   = float(close.tail(10).mean()) if len(close)  >= 1 else 0.0
                data["vol_5"]   = int(volume.tail(5).mean())   if len(volume) >= 1 else 0

                # 最終防線：任一關鍵欄位仍為 NaN（理論上不應發生）→ 跳過
                import math
                if any(math.isnan(v) for v in [data["price"], data["high_20"], data["ma_10"]]):
                    data = {"price": 0.0, "high": 0.0, "low": 0.0, "vol": 0,
                            "high_20": 0.0, "ma_10": 0.0, "vol_5": 0}
                    continue

                with _CACHE_LOCK:
                    TECH_DATA_CACHE[sid] = data
                return data  # 任一 suffix 成功 → 立即結束，不浪費時間

            except Exception:
                pass  # 單次錯誤，繼續試下一個 suffix

        # 同一 attempt 的兩個 suffix 都失敗（真正網路異常），才 sleep
        if attempt < YF_RETRIES:
            time.sleep(YF_BACKOFF ** attempt)

    # 所有 attempt 全失敗，寫入空值快取避免重複卡頓

    with _CACHE_LOCK:
        TECH_DATA_CACHE[sid] = data
    return data


def get_taiex_above_ma20() -> bool:
    """
    大盤是否站上月線（20日均線）。
    用 ^TWII（加權指數）。快取一份即可。
    """
    cache_key = "taiex"
    if cache_key in TAIEX_MA20_CACHE:
        return TAIEX_MA20_CACHE[cache_key]

    result = False
    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
            # 同樣使用 3mo，確保 tail(20) 拿到真實 20 個交易日均線
            hist = yf.Ticker("^TWII").history(period="3mo")
        if hist is not None and not hist.empty:
            last_close = float(hist["Close"].iloc[-1])
            ma20       = float(hist["Close"].tail(20).mean())
            result     = last_close > ma20
    except Exception:
        pass

    TAIEX_MA20_CACHE[cache_key] = result
    return result


# ─────────────────────────────────────────────────────────────────────────────
#  優化一：多執行緒預載（Multithreading Prefetch）
# ─────────────────────────────────────────────────────────────────────────────
def prefetch_tech_data(sids: list[str]) -> None:
    """
    用 ThreadPoolExecutor 同時抓取所有股票的技術面資料，
    結果寫入 TECH_DATA_CACHE（有 _CACHE_LOCK 保護）。

    串行抓 100 檔約需 3~5 分鐘；並行後縮短到 10~20 秒。
    TECH_WORKERS 預設 10，可透過環境變數調整（建議 8~12）。
    """
    # 過濾掉已快取的，避免重複請求
    to_fetch = [s for s in sids if s not in TECH_DATA_CACHE]
    if not to_fetch:
        return

    print(f"[INFO] 並行抓取技術面資料：{len(to_fetch)} 檔（workers={TECH_WORKERS}）…")

    def _fetch(sid: str):
        try:
            get_stock_tech_data(sid)   # 內部已處理快取 + lock
        except Exception as e:
            print(f"[WARN] prefetch {sid} 失敗：{e}")

    with ThreadPoolExecutor(max_workers=TECH_WORKERS) as pool:
        futures = {pool.submit(_fetch, sid): sid for sid in to_fetch}
        done = 0
        for fut in as_completed(futures):
            done += 1
            if done % 20 == 0 or done == len(to_fetch):
                print(f"[INFO]   技術面預載進度：{done}/{len(to_fetch)}")

    print(f"[INFO] 技術面預載完成，快取筆數：{len(TECH_DATA_CACHE)}")


# ─────────────────────────────────────────────────────────────────────────────
#  半量化訊號計算
# ─────────────────────────────────────────────────────────────────────────────
def calc_signals(tech: dict, net_qty: int, taiex_above_ma20: bool) -> dict:
    """
    依 Word 文件「半量化交易系統」計算所有訊號欄位。

    進場價  = MAX(今日收盤, 20日高點) × 1.01
    停損價  = 進場價 × 0.92
    停利1   = 進場價 × 1.20
    停利2   = 進場價 × 1.35
    移動停損 = 10日均線

    突破20日高 = 今日最高價 > 20日高點
    放量1.5倍  = 今日成交量 > 5日均量 × 1.5
    大盤站上月線 = ^TWII > 大盤20日均
    綜合判斷   = 突破 AND 放量 AND 大盤 AND 外資淨買超 → 🔥可買
    """
    price   = tech["price"]
    high    = tech["high"]
    high_20 = tech["high_20"]
    ma_10   = tech["ma_10"]
    vol     = tech["vol"]
    vol_5   = tech["vol_5"]

    # ── 進場/停損/停利 ────────────────────────────────────────────────────────
    entry = round(max(price, high_20) * 1.01, 2) if high_20 > 0 else 0.0
    stop  = round(entry * 0.92, 2) if entry > 0 else 0.0
    tp1   = round(entry * 1.20, 2) if entry > 0 else 0.0
    tp2   = round(entry * 1.35, 2) if entry > 0 else 0.0
    trail = round(ma_10, 2) if ma_10 > 0 else 0.0

    # ── 訊號判斷 ──────────────────────────────────────────────────────────────
    is_breakout  = (high > high_20) if (high > 0 and high_20 > 0) else False
    is_high_vol  = (vol > vol_5 * 1.5) if (vol > 0 and vol_5 > 0) else False
    is_net_buy   = net_qty > 0

    # 優化二：空頭防禦 — 股價必須站上 10 日均線
    # 若外資買超但股價仍在均線下方，通常是「逢低攤平」或「借券回補」，
    # 非真正多頭啟動，此時進場勝率偏低，過濾掉。
    is_above_ma10 = (price >= ma_10) if (price > 0 and ma_10 > 0) else False

    # 流動性濾網：5日均量必須達到最低門檻（預設 500 張 = 500,000 股）
    # 冷門股 / 殭屍股今天偶爾爆量仍可觸發「放量1.5倍」，但買進賣出都容易滑價，
    # 甚至想停損賣不掉。設定基礎流動性門檻可有效排除這類陷阱。
    # 可透過環境變數 MIN_VOL5 調整（單位：張，1張=1000股）
    MIN_VOL5_LOTS = int(os.getenv("MIN_VOL5", "500"))   # 預設 500 張
    MIN_VOL5_SHARES = MIN_VOL5_LOTS * 1000              # 換算為股數
    is_liquid = (vol_5 >= MIN_VOL5_SHARES) if vol_5 > 0 else False

    # 綜合判斷：突破 + 放量 + 流動性 + 大盤站月線 + 外資淨買超 + 站上10均
    all_pass = (is_breakout and is_high_vol and is_liquid
                and taiex_above_ma20 and is_net_buy and is_above_ma10)
    verdict  = "🔥可買" if all_pass else "觀察"

    # 優化三：風報比（Risk / Reward Ratio）
    # 每張風險 = (進場價 - 停損價) × 1000 股，讓操盤手直接看 NT$ 曝險
    risk_per_lot   = round((entry - stop) * 1000, 0) if entry > stop else 0.0
    reward_tp1_lot = round((tp1   - entry) * 1000, 0) if tp1 > entry else 0.0
    reward_tp2_lot = round((tp2   - entry) * 1000, 0) if tp2 > entry else 0.0
    # R/R 比 = 潛在獲利 / 潛在損失（以 TP1 為基準）；固定公式下恆為 2.5
    rr_ratio = round(reward_tp1_lot / risk_per_lot, 2) if risk_per_lot > 0 else 0.0

    return {
        "進場價":       entry,
        "停損價":       stop,
        "停利1":        tp1,
        "停利2":        tp2,
        "移動停損":     trail,
        "每張風險NT$":  int(risk_per_lot),
        "每張獲利NT$":  int(reward_tp1_lot),
        "風報比":       rr_ratio,
        "突破20日高":   "✅" if is_breakout      else "❌",
        "放量1.5倍":    "✅" if is_high_vol      else "❌",
        "流動性足夠":   "✅" if is_liquid        else "❌",
        "站上10均":     "✅" if is_above_ma10    else "❌",
        "大盤站月線":   "✅" if taiex_above_ma20 else "❌",
        "綜合判斷":     verdict,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  爬蟲 / 解析
# ─────────────────────────────────────────────────────────────────────────────
def try_fetch_and_parse(a_code: str, b_code: str, days: int):
    url_qty = f"{BASE_URL}?a={a_code}&b={b_code}&c=E&d={days}"
    url_amt = f"{BASE_URL}?a={a_code}&b={b_code}&c=B&d={days}"
    html_qty = fetch_html(url_qty)
    html_amt = fetch_html(url_amt)
    dump_debug_html(f"{a_code}_{b_code}_E", html_qty)
    dump_debug_html(f"{a_code}_{b_code}_B", html_amt)
    if (NO_DATA_TEXT in html_qty) and (NO_DATA_TEXT in html_amt):
        return {}, {}, url_qty, url_amt, True
    return parse_table(html_qty), parse_table(html_amt), url_qty, url_amt, False


# ─────────────────────────────────────────────────────────────────────────────
#  主報表建立
# ─────────────────────────────────────────────────────────────────────────────
REPORT_COLUMNS = [
    "日期", "代碼", "名稱", "大戶",
    "買進", "賣出", "淨超",
    "區間均價", "現價", "乖離率",
    # 技術面原始數據
    "20日高", "10日均", "5日均量", "今日高", "今日量",
    # 價格訊號
    "進場價", "停損價", "停利1", "停利2", "移動停損",
    # 風報比（優化三）
    "每張風險NT$", "每張獲利NT$", "風報比",
    # 布林訊號
    "突破20日高", "放量1.5倍", "流動性足夠", "站上10均", "大盤站月線",
    "綜合判斷",
]

FAIL_COLUMNS = [
    "日期", "券商key", "總公司(a)", "嘗試分點(b)",
    "券商名稱", "錯誤訊息", "網址(張數E)", "網址(金額B)",
]


def build_report(days: int):
    start_time = now_taipei()
    rows, failures, errors = [], [], []
    broker_ok = broker_fail = 0

    # 大盤月線（一次查好，全部股票共用）
    print("[INFO] 查詢大盤月線（^TWII）…")
    taiex_above_ma20 = get_taiex_above_ma20()
    taiex_label = "✅站上" if taiex_above_ma20 else "❌未站上"
    print(f"[INFO] 大盤站上月線：{taiex_label}")

    # ── 優化一：收集所有外資出現的股票代碼，並行預載技術面資料 ────────────────
    # 先用一輪爬蟲把所有 sid 收集齊，再一次性並行拉 yfinance，
    # 大幅縮短從「逐檔串行」到「全部並行」的等待時間。
    print("[INFO] 預掃描各外資持股清單（用於並行技術面預載）…")
    all_sids: set[str] = set()
    broker_cache: dict = {}   # 暫存爬蟲結果，避免後續重複爬取

    for bk, (ac, _) in LEGEND_BROKERS.items():
        ac = str(ac)
        for b_try in [ac, "9900"]:
            try:
                q_map, a_map, u_qty, u_amt, no_data = try_fetch_and_parse(ac, b_try, days)
                if no_data:
                    continue
                if q_map or a_map:
                    common = set(q_map.keys()) & set(a_map.keys())
                    all_sids.update(common)
                    broker_cache[bk] = (ac, q_map, a_map, u_qty, u_amt)
                    break
            except Exception:
                pass

    # 並行預載全部技術面（結果存入 TECH_DATA_CACHE）
    prefetch_tech_data(list(all_sids))

    for broker_key, (a_code, broker_label) in LEGEND_BROKERS.items():
        a_code = str(a_code)
        candidate_bs = [a_code, "9900"]
        chosen_b = None
        qty_map = amt_map = {}
        url_qty = url_amt = ""

        try:
            # 優先使用預掃描已快取的爬蟲結果（避免重複 HTTP 請求）
            if broker_key in broker_cache:
                _ac, qty_map, amt_map, url_qty, url_amt = broker_cache[broker_key]
                chosen_b = _ac
            else:
                for b_try in candidate_bs:
                    q_map, a_map, u_qty, u_amt, no_data = try_fetch_and_parse(a_code, b_try, days)
                    url_qty, url_amt = u_qty, u_amt
                    if no_data:
                        continue
                    if q_map or a_map:
                        chosen_b = b_try
                        qty_map, amt_map = q_map, a_map
                        break

            if not chosen_b:
                raise RuntimeError(f"查詢結果：{NO_DATA_TEXT}（已嘗試 b=a 與 b=9900）")

            common = set(qty_map.keys()) & set(amt_map.keys())
            if not common:
                raise RuntimeError("E/B 解析後無交集")

            for sid in common:
                info    = qty_map[sid]
                net_qty = info["net"]
                net_amt = amt_map[sid]["net"]
                avg_cost = round((net_amt / net_qty), 2) if net_qty > 0 else 0.0

                # ── 技術面 ────────────────────────────────────────────────────
                tech  = get_stock_tech_data(sid)
                price = tech["price"]
                bias  = (
                    round(((price - avg_cost) / avg_cost) * 100, 2)
                    if (price > 0 and avg_cost > 0)
                    else 0.0
                )

                # ── 半量化訊號 ────────────────────────────────────────────────
                sig = calc_signals(tech, net_qty, taiex_above_ma20)

                rows.append({
                    "日期":     start_time.strftime("%Y%m%d"),
                    "代碼":     sid,
                    "名稱":     info["name"],
                    "大戶":     broker_label,
                    "買進":     info["buy"],
                    "賣出":     info["sell"],
                    "淨超":     int(net_qty),
                    "區間均價": avg_cost,
                    "現價":     round(price, 2),
                    "乖離率":   f"{bias}%",
                    # 技術面原始數據
                    "20日高":   round(tech["high_20"], 2),
                    "10日均":   round(tech["ma_10"], 2),
                    "5日均量":  tech["vol_5"],
                    "今日高":   round(tech["high"], 2),
                    "今日量":   tech["vol"],
                    # 半量化訊號
                    **sig,
                })

            broker_ok += 1
            time.sleep(1.2)

        except Exception as e:
            broker_fail += 1
            errors.append(f"{broker_label}({broker_key}) 失敗：{e}")
            failures.append({
                "日期":       start_time.strftime("%Y%m%d"),
                "券商key":    str(broker_key),
                "總公司(a)":  a_code,
                "嘗試分點(b)": ",".join(candidate_bs),
                "券商名稱":   broker_label,
                "錯誤訊息":   str(e),
                "網址(張數E)": url_qty,
                "網址(金額B)": url_amt,
            })

    df      = pd.DataFrame(rows,     columns=REPORT_COLUMNS)
    fail_df = pd.DataFrame(failures, columns=FAIL_COLUMNS)

    summary = {
        "generated_at":  start_time.isoformat(),
        "timezone":      "Asia/Taipei",
        "days":          days,
        "total_rows":    int(len(df)),
        "brokers_total": int(len(LEGEND_BROKERS)),
        "brokers_ok":    broker_ok,
        "brokers_fail":  broker_fail,
        "success":       (broker_fail == 0),
        "errors":        errors[:50],
        "taiex_above_ma20": taiex_above_ma20,
    }

    # ── 外資排序 + 各家內部淨超排序 ──────────────────────────────────────────
    if not df.empty:
        df["淨超"] = pd.to_numeric(df["淨超"], errors="coerce").fillna(0).astype(int)
        broker_order = (
            df.groupby("大戶")["淨超"]
              .sum()
              .sort_values(ascending=False)
              .index.tolist()
        )
        df["大戶"] = pd.Categorical(df["大戶"], categories=broker_order, ordered=True)
        df = df.sort_values(["大戶", "淨超"], ascending=[True, False]).reset_index(drop=True)

    # ── summary top_preview ────────────────────────────────────────────────
    TOP_N = int(os.getenv("TOP_N", "10"))
    top_preview = []

    if not df.empty and hasattr(df["大戶"], "cat"):
        for broker in df["大戶"].cat.categories.tolist():
            sub = df[df["大戶"] == broker].head(TOP_N)
            if sub.empty:
                continue
            top_preview.append({
                "broker":    str(broker),
                "total_net": int(df[df["大戶"] == broker]["淨超"].sum()),
                "rows": [
                    {
                        "sid":     r["代碼"],
                        "name":    r["名稱"],
                        "net":     int(r["淨超"]),
                        "avg":     r["區間均價"],
                        "price":   r["現價"],
                        "bias":    r["乖離率"],
                        "entry":   r["進場價"],
                        "stop":    r["停損價"],
                        "tp1":     r["停利1"],
                        "tp2":     r["停利2"],
                        "verdict": r["綜合判斷"],
                    }
                    for _, r in sub.iterrows()
                ],
            })

    summary["top_preview"] = top_preview
    summary["top_n"]       = TOP_N

    if summary["total_rows"] == 0:
        summary["success"] = False
        summary["errors"]  = (summary.get("errors") or []) + [
            "總資料筆數為 0（全部查無分點交易資料或解析規則需調整）"
        ]

    return df, fail_df, summary


# ─────────────────────────────────────────────────────────────────────────────
#  Excel 輸出
# ─────────────────────────────────────────────────────────────────────────────
# REPORT_COLUMNS（29欄）完整對照表（A → AC）
# A  日期      B  代碼      C  名稱      D  大戶
# E  買進      F  賣出      G  淨超      H  區間均價   I  現價      J  乖離率
# K  20日高    L  10日均    M  5日均量   N  今日高     O  今日量
# P  進場價    Q  停損價    R  停利1     S  停利2      T  移動停損
# U  每張風險NT$  V  每張獲利NT$  W  風報比
# X  突破20日高  Y  放量1.5倍  Z  流動性足夠  AA  站上10均  AB  大盤站月線  AC  綜合判斷
_COL_WIDTHS = {
    "A": 10, "B":  8, "C": 14, "D": 35,        # 日期/代碼/名稱/大戶
    "E": 10, "F": 10, "G": 10, "H": 12,        # 買進/賣出/淨超/區間均價
    "I": 10, "J": 10,                            # 現價/乖離率
    "K": 10, "L": 10, "M": 10, "N": 10, "O": 10,  # 技術面五欄
    "P": 10, "Q": 10, "R": 10, "S": 10, "T": 10,  # 訊號價格五欄
    "U": 13, "V": 13, "W": 10,                  # 風報比三欄
    "X": 12, "Y": 12, "Z": 12, "AA": 12, "AB": 12, "AC": 15,  # 布林訊號(5) + 綜合判斷
}


def export_excel(df: pd.DataFrame, fail_df: pd.DataFrame, xlsx_path: str):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import FormulaRule

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        # ── Report sheet ──────────────────────────────────────────────────────
        df.to_excel(writer, index=False, sheet_name="Report")
        ws = writer.sheets["Report"]

        for col, w in _COL_WIDTHS.items():
            ws.column_dimensions[col].width = w

        last_row = ws.max_row
        if last_row >= 2:
            # 乖離率（J欄）條件上色
            rng_bias = f"J2:J{last_row}"
            fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_blue  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            fill_red   = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)

            ws.conditional_formatting.add(rng_bias, FormulaRule(
                formula=[r'=VALUE(SUBSTITUTE($J2,"%",""))>10'],
                fill=fill_red, font=font_white, stopIfTrue=True))
            ws.conditional_formatting.add(rng_bias, FormulaRule(
                formula=[r'=AND(VALUE(SUBSTITUTE($J2,"%",""))>3,VALUE(SUBSTITUTE($J2,"%",""))<=10)'],
                fill=fill_blue, stopIfTrue=True))
            ws.conditional_formatting.add(rng_bias, FormulaRule(
                formula=[r'=VALUE(SUBSTITUTE($J2,"%",""))<=3'],
                fill=fill_green, stopIfTrue=True))

            from openpyxl.utils import get_column_letter

            # 綜合判斷欄：🔥可買 → 金黃底
            verdict_col = get_column_letter(REPORT_COLUMNS.index("綜合判斷") + 1)
            rng_verdict = f"{verdict_col}2:{verdict_col}{last_row}"
            fill_gold = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            font_bold = Font(bold=True)
            ws.conditional_formatting.add(rng_verdict, FormulaRule(
                formula=[f'=${verdict_col}2="🔥可買"'],
                fill=fill_gold, font=font_bold, stopIfTrue=True))

            # 優化三：風報比（W欄）條件上色
            # RR ≥ 2.5（TP1 目標達成）→ 綠底；< 2.0 → 紅底（勿輕易進場）
            rr_col = get_column_letter(REPORT_COLUMNS.index("風報比") + 1)
            rng_rr = f"{rr_col}2:{rr_col}{last_row}"
            fill_rr_good = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_rr_bad  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            ws.conditional_formatting.add(rng_rr, FormulaRule(
                formula=[f'=${rr_col}2>=2.5'],
                fill=fill_rr_good, stopIfTrue=True))
            ws.conditional_formatting.add(rng_rr, FormulaRule(
                formula=[f'=AND(${rr_col}2>0,${rr_col}2<2)'],
                fill=fill_rr_bad, stopIfTrue=True))

            # 站上10均（空頭防禦欄）：❌ → 橙底警示
            ma10_col = get_column_letter(REPORT_COLUMNS.index("站上10均") + 1)
            rng_ma10 = f"{ma10_col}2:{ma10_col}{last_row}"
            fill_warn = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            ws.conditional_formatting.add(rng_ma10, FormulaRule(
                formula=[f'=${ma10_col}2="❌"'],
                fill=fill_warn, stopIfTrue=True))

        # ── TopByBroker sheet ─────────────────────────────────────────────────
        top_n = int(os.getenv("TOP_N", "10"))
        if df is not None and not df.empty:
            top_df = df.groupby("大戶", sort=False).head(top_n)
            top_df.to_excel(writer, index=False, sheet_name="TopByBroker")
            ws3 = writer.sheets["TopByBroker"]
            for col, w in _COL_WIDTHS.items():
                ws3.column_dimensions[col].width = w

        # ── 訊號篩選 sheet（只列出 🔥可買）────────────────────────────────────
        if df is not None and not df.empty:
            buy_df = df[df["綜合判斷"] == "🔥可買"].copy()
            if not buy_df.empty:
                buy_df.to_excel(writer, index=False, sheet_name="可買訊號")
                ws_buy = writer.sheets["可買訊號"]
                for col, w in _COL_WIDTHS.items():
                    ws_buy.column_dimensions[col].width = w

        # ── Failures sheet ────────────────────────────────────────────────────
        if fail_df is not None and not fail_df.empty:
            fail_df.to_excel(writer, index=False, sheet_name="Failures")
            ws2 = writer.sheets["Failures"]
            for col, w in zip("ABCDEFGH", [10,12,10,14,35,60,55,55]):
                ws2.column_dimensions[col].width = w


# ─────────────────────────────────────────────────────────────────────────────
#  PDF 輸出
# ─────────────────────────────────────────────────────────────────────────────
def export_pdf(df: pd.DataFrame, pdf_path: str, summary: dict):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_path = os.path.join("fonts", "NotoSansTC-Regular.ttf")
    font_name = "NotoSansTC"

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        leftMargin=12, rightMargin=12, topMargin=14, bottomMargin=14,
    )
    styles    = getSampleStyleSheet()
    normal    = styles["Normal"]
    title_sty = styles["Title"]

    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont(font_name, font_path))
        normal.fontName = title_sty.fontName = font_name
    else:
        normal.fontName = title_sty.fontName = "Helvetica"

    taiex_lbl = "✅站上月線" if summary.get("taiex_above_ma20") else "❌未站上月線"
    title_para = Paragraph("【外資分點狙擊 × 半量化交易訊號】報表", title_sty)
    meta_para  = Paragraph(
        f"產生時間：{summary['generated_at']}（{summary['timezone']}）　"
        f"查詢區間：{summary['days']} 日　"
        f"大盤月線：{taiex_lbl}　"
        f"筆數：{summary['total_rows']}　"
        f"結果：{'成功' if summary['success'] else '失敗/部分失敗'}",
        normal,
    )
    elements = [title_para, Spacer(1, 6), meta_para, Spacer(1, 10)]

    # ── 核心欄位（PDF 顯示精簡版，省略 5日均量/今日高/今日量）────────────────
    PDF_COLS = [
        "代碼", "名稱", "大戶", "淨超", "區間均價", "現價", "乖離率",
        "20日高", "10日均",
        "進場價", "停損價", "停利1", "停利2",
        "每張風險NT$", "風報比",
        "突破20日高", "放量1.5倍", "流動性足夠", "站上10均", "大盤站月線", "綜合判斷",
    ]

    if df is not None and not df.empty:
        pdf_df  = df[PDF_COLS]
        header  = PDF_COLS
        data    = [header] + pdf_df.astype(str).values.tolist()
    else:
        data = [PDF_COLS]

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME",    (0, 0), (-1, -1), normal.fontName),
        ("FONTSIZE",    (0, 0), (-1, -1), 7),
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#EAEAEA")),
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
    ]))
    elements.append(tbl)

    if summary.get("errors"):
        elements += [Spacer(1, 10), Paragraph("錯誤摘要（前 20 筆）：", normal)]
        for e in summary["errors"][:20]:
            elements.append(Paragraph(f"- {e}", normal))

    doc.build(elements)


# ─────────────────────────────────────────────────────────────────────────────
#  主程式
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ensure_output_dir()
    days = int(os.getenv("DAYS", str(DEFAULT_DAYS)))

    df, fail_df, summary = build_report(days)

    ymd          = now_taipei().strftime("%Y%m%d")
    xlsx_path    = os.path.join("output", f"IKE_Report_{ymd}.xlsx")
    pdf_path     = os.path.join("output", f"IKE_Report_{ymd}.pdf")
    summary_path = os.path.join("output", "summary.json")

    export_excel(df, fail_df, xlsx_path)
    export_pdf(df, pdf_path, summary)

    # ── 大盤籌碼 ──────────────────────────────────────────────────────────────
    if HAS_MARKET_DATA and os.getenv("MARKET_DATA", "1") == "1":
        try:
            top_stock_ids, seen = [], set()
            for block in (summary.get("top_preview") or [])[:5]:
                for r in (block.get("rows") or [])[:3]:
                    sid = r.get("sid", "")
                    if sid and sid not in seen:
                        top_stock_ids.append(sid)
                        seen.add(sid)

            market = fetch_all_market_data(
                top_stock_ids=top_stock_ids[:10],
                history_days=int(os.getenv("MARKET_HISTORY_DAYS", "6")),
            )
            summary["market_data"] = market
            if market.get("fetch_errors"):
                summary["errors"] = (summary.get("errors") or []) + [
                    f"[market] {e}" for e in market["fetch_errors"]
                ]
            print(f"[OK] 大盤籌碼: inst={len(market.get('institutional', []))}, "
                  f"taiex={len(market.get('taiex', []))}")
        except Exception as e:
            print(f"[WARN] 大盤籌碼抓取失敗（非致命）: {e}")
            summary["market_data"] = {"fetch_errors": [str(e)]}
    else:
        print("[INFO] 大盤籌碼抓取已停用")

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f"[OK] Excel  : {xlsx_path}")
    print(f"[OK] PDF    : {pdf_path}")
    print(f"[OK] Summary: {summary_path}")

    # ── 訊號摘要 ──────────────────────────────────────────────────────────────
    if not df.empty:
        buy_signals = df[df["綜合判斷"] == "🔥可買"]
        if not buy_signals.empty:
            print(f"\n🔥 共 {len(buy_signals)} 檔符合進場條件：")
            for _, r in buy_signals.iterrows():
                print(f"   [{r['代碼']}] {r['名稱']}  進場:{r['進場價']}  "
                      f"停損:{r['停損價']}  停利1:{r['停利1']}  停利2:{r['停利2']}")
        else:
            print("\n[INFO] 今日無符合半量化進場條件的個股。")

    if summary.get("total_rows", 0) == 0 or summary.get("brokers_ok", 0) == 0:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
